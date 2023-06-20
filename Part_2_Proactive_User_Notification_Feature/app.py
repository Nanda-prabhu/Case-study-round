import os
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from openpyxl import Workbook, load_workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

linkedin_username = "nanda9******"
linkedin_password = "GF4C******"

recipient_email = "nanda96here@gmail.com"

excel_file = "linkedin_data.xlsx"

driver = webdriver.Edge()

def login_to_linkedin(username, password):
    driver.get("https://www.linkedin.com/login")
    time.sleep(2)

    username_field = driver.find_element(By.ID, "username")
    username_field.send_keys(username)
    password_field = driver.find_element(By.ID, "password")
    password_field.send_keys(password)
    password_field.send_keys(Keys.RETURN)

    try:
        WebDriverWait(driver, 10).until(EC.url_contains("https://www.linkedin.com/feed/"))
    except TimeoutException:
        print("Login failed. Please check your credentials.")
        driver.quit()
        exit()

def get_unread_data():
    driver.get("https://www.linkedin.com/feed/")
    driver.implicitly_wait(10)  # Adjust the wait time as needed
    
    navbar = driver.find_element(By.CSS_SELECTOR, ".global-nav__content")
    unread_counts = navbar.find_elements(By.CSS_SELECTOR, ".notification-badge__count")
    
    message_count = int(unread_counts[0].text)
    notification_count = int(unread_counts[1].text)

    return message_count, notification_count


def save_data_to_excel(messages, notifications):
    if not os.path.exists(excel_file):
        workbook = Workbook()
        sheet = workbook.active
        sheet.cell(row=1, column=1, value="Unread Messages")
        sheet.cell(row=1, column=2, value="Unread Notifications")
    else:
        workbook = load_workbook(excel_file)
        sheet = workbook.active

    last_row = sheet.max_row
    sheet.cell(row=last_row+1, column=1, value=messages)
    sheet.cell(row=last_row+1, column=2, value=notifications)
    workbook.save(excel_file)

def send_email_notification(messages, notifications, prev_messages, prev_notifications):
    # Email configuration
    smtp_server = "smtp.gmail.com"
    smtp_port = 587
    sender_email = "nanda96*******"
    sender_password = "mrto********"
    recipient_email = "nanda96here@gmail.com"

    email_subject = "LinkedIn Unread Notifications"

    # HTML formatting for the email body
    email_body = f"""
        <html>
        <head>
            <style>
            .card {{
                font-family: Verdana;
                background-color: whitesmoke;
                width: max-content;
                padding: 1rem;
                border-radius: 20px;
                box-shadow: 0px 10px 15px -3px rgba(0,0,0,0.1),0px 10px 15px -3px rgba(0,0,0,0.1),0px 10px 15px -3px rgba(0,0,0,0.1),0px 10px 15px -3px rgba(0,0,0,0.1);
            }}
            p {{
                margin-bottom: 10px;
            }}
            .highlight {{
                font-weight: bold;
                color: #007bff;
            }}
            </style>
        </head>
        <body>
            <div class="card">
            <h1>LinkedIn Unread Notifications</h1>
            <p><span style="font-weight: bold; color: #007bff;">Number of Unread Messages:</span> {messages}</p>
            <p><span style="font-weight: bold; color: #007bff;">Number of Unread Notifications:</span> {notifications}</p>
            <h2>Comparison with previous occurrence:</h2>
            <p><span style="font-weight: bold; color: #007bff;">Unread Messages Change:</span> {messages - prev_messages}</p>
            <p><span style="font-weight: bold; color: #007bff;">Unread Notifications Change:</span> {notifications - prev_notifications}</p>
            </div>
        </body>
        </html>

    """

    message = MIMEMultipart()
    message["From"] = sender_email
    message["To"] = recipient_email
    message["Subject"] = email_subject
    message.attach(MIMEText(email_body, "html"))

    try:
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(sender_email, sender_password)
        server.send_message(message)
        server.quit()
        print("Email sent successfully :)")
    except smtplib.SMTPException as e:
        print("Error sending email notification:", str(e))


def main():
    login_to_linkedin(linkedin_username, linkedin_password)
    current_messages, current_notifications = get_unread_data()

    if os.path.exists(excel_file):
        workbook = load_workbook(excel_file)
        sheet = workbook.active
        prev_row = sheet.max_row - 1
        prev_messages = sheet.cell(row=prev_row, column=1).value
        prev_notifications = sheet.cell(row=prev_row, column=2).value

    save_data_to_excel(current_messages, current_notifications)
    # send_email_notification(current_messages, current_notifications)
    send_email_notification(current_messages, current_notifications, prev_messages, prev_notifications)


if __name__ == "__main__":
    main()
